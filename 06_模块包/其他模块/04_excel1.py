from openpyxl import Workbook
from openpyxl import load_workbook

# 获取单元格的值
wb = load_workbook("住户模板.xlsx")
sheet = wb.get_sheet_by_name("住户模板") # 通过sheet名字获取excel中的对象
print(sheet["A1"].value) # 获取一个excel中的表格中数据
print(sheet["A3:D3"]) # 获取指定位置excel中的表格中数据

# 循环获取所有的数据
for row in sheet:
    # print(row)
    for cell in row:
        print(cell.value, end=",")
    print("\n")
print("\n")
print("\n")

# 按行循环打印指定数据
for row in sheet.iter_rows(min_row=3, max_row=12, max_col=8):
    for cell in row:
        print(cell.value, end=",")
    print("\n")

print("\n")
print("\n")

# 按列循环打印指定数据
for cell in sheet.iter_cols(min_col=1, max_col=4, min_row=3, max_row=12):
    # print(cell, 'cell')
    for row in cell:
        print(row.value, end=",")
    print("\n")

sheet["A3"].value = "李四"
print(sheet["A3"].value)
wb.save("住户模板.xlsx")

# 删除工作表
# wb.remove(sheet)